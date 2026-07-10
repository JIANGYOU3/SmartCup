"""
按发布时间拆分为近三年 / 三年前两份文件，每份含汇总+数据明细

用法：
  conda run -n SmartCup python source/cleandata/split_by_time.py
"""

import csv
import sys
import re
from pathlib import Path
from collections import Counter
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))
from source.common.excel_style import csv_to_excel

DATA = Path(__file__).resolve().parent.parent.parent / "res/data/douyin/output"
INPUT = DATA / "标签结果_最终.csv"
CUTOFF = datetime(2023, 7, 1)

# ── 样式 ──
title_font = Font(name="微软雅黑", size=14, bold=True, color="1F4E79")
h2_font = Font(name="微软雅黑", size=12, bold=True, color="2E75B6")
header_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="2E75B6")
body_font = Font(name="微软雅黑", size=10)
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                     top=Side(style="thin"), bottom=Side(style="thin"))
center = Alignment(horizontal="center", vertical="center")
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)


def safe_int(v):
    try: return int(v or 0)
    except: return 0


def parse_date(r):
    t = r.get("发布时间", "").strip()
    try: return datetime.strptime(t[:10], '%Y-%m-%d')
    except: return None


def write_summary_sheet(ws, rows, label):
    """写汇总 sheet"""
    N = len(rows)
    total_likes = sum(safe_int(r.get("点赞数", 0)) for r in rows)
    total_comments = sum(safe_int(r.get("评论数", 0)) for r in rows)
    total_collects = sum(safe_int(r.get("收藏数", 0)) for r in rows)
    total_shares = sum(safe_int(r.get("分享数", 0)) for r in rows)
    total_followers = sum(safe_int(r.get("作者粉丝量", 0)) for r in rows)

    # 统计
    label_count = Counter()
    emotion_count = Counter()
    topic_count = Counter()
    value_count = Counter()
    relevance_count = Counter()
    pain_points = Counter()
    all_keywords = Counter()

    for r in rows:
        for lb in r.get("内容标签", "").replace("，", ",").split(","):
            lb = lb.strip().strip("[]'\" ")
            if lb: label_count[lb] += 1
        emo = r.get("用户情绪", "").strip()
        if emo: emotion_count[emo] += 1
        tp = r.get("话题类别", "").strip()
        if tp: topic_count[tp] += 1
        vl = r.get("数据价值等级", "").strip()
        if vl: value_count[vl] += 1
        rl = r.get("相关性等级", "").strip()
        if rl: relevance_count[rl] += 1
        pains = r.get("需求痛点", "").strip()
        if pains and pains != "无":
            for p in pains.replace("；", ";").replace("，", ",").split(";"):
                p = p.strip()
                if p and len(p) > 2: pain_points[p[:40]] += 1
        kws = r.get("关键词", "").strip()
        if kws:
            for kw in kws.replace("；", ";").replace("，", ",").split(","):
                kw = kw.strip().strip("[]'\" ")
                if kw: all_keywords[kw] += 1

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16

    row = 1

    # 标题
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    c = ws.cell(row=row, column=1, value=f"📊 抖音智能水杯数据汇总 — {label}")
    c.font = title_font
    row += 2

    # 总览
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row=row, column=1, value="📋 数据总览").font = h2_font
    row += 1
    overview = [
        ["#", "指标", "数值", "占比", "备注"],
        [1, "视频总数", f"{N:,}", "100%", ""],
        [2, "总点赞数", f"{total_likes:,}", "-", f"均值 {total_likes//N:,}/条"],
        [3, "总评论数", f"{total_comments:,}", "-", f"均值 {total_comments//N:,}/条"],
        [4, "总收藏数", f"{total_collects:,}", "-", f"均值 {total_collects//N:,}/条"],
        [5, "总分享数", f"{total_shares:,}", "-", f"均值 {total_shares//N:,}/条"],
        [6, "作者粉丝总量", f"{total_followers:,}", "-", "-"],
    ]
    for dr in overview:
        for ci, val in enumerate(dr):
            cell = ws.cell(row=row, column=ci+1, value=val)
            cell.font = header_font if dr == overview[0] else body_font
            cell.fill = header_fill if dr == overview[0] else PatternFill()
            cell.alignment = left_align if ci == 1 else center
            cell.border = thin_border
        row += 1
    row += 1

    def write_section(title_text, items, top_n=15):
        nonlocal row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.cell(row=row, column=1, value=title_text).font = h2_font
        row += 1
        hdr = ["#", "名称", "数量", "占比", ""]
        for ci, v in enumerate(hdr):
            cell = ws.cell(row=row, column=ci+1, value=v)
            cell.font = header_font; cell.fill = header_fill
            cell.alignment = center; cell.border = thin_border
        row += 1
        for i, (name, cnt) in enumerate(items[:top_n]):
            for ci, v in enumerate([i+1, name, cnt, f"{cnt/N*100:.1f}%", ""]):
                cell = ws.cell(row=row, column=ci+1, value=v)
                cell.font = body_font
                cell.alignment = left_align if ci == 1 else center
                cell.border = thin_border
            row += 1
        row += 1

    write_section("🏷️ 内容标签分布", label_count.most_common(20))
    write_section("📂 话题类别", topic_count.most_common(15))
    write_section("😊 用户情绪", emotion_count.most_common())
    write_section("🎯 相关性等级", relevance_count.most_common())
    write_section("💎 数据价值等级", value_count.most_common())
    write_section("🔍 用户需求痛点 Top 20", pain_points.most_common(20))

    # 关键词
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row=row, column=1, value="🔑 AI提取关键词 Top 25").font = h2_font
    row += 1
    top_kw = all_keywords.most_common(25)
    for i in range(0, len(top_kw), 5):
        chunk = top_kw[i:i+5]
        for ci, (kw, cnt) in enumerate(chunk):
            cell = ws.cell(row=row, column=ci+1, value=f"{kw}({cnt})")
            cell.font = body_font; cell.alignment = center; cell.border = thin_border
        row += 1

    ws.freeze_panes = "A2"


def main():
    # 读取
    with open(INPUT, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        all_rows = list(reader)
        fieldnames = reader.fieldnames

    # 拆分
    recent, older = [], []
    for r in all_rows:
        dt = parse_date(r)
        if dt and dt >= CUTOFF:
            recent.append(r)
        else:
            older.append(r)

    print(f"近三年 (≥2023-07): {len(recent)} 条")
    print(f"三年前  (<2023-07): {len(older)} 条")

    for label, rows, prefix in [
        ("近三年", recent, "近三年"),
        ("三年前", older, "三年前"),
    ]:
        if not rows:
            print(f"  ⚠️ {label}: 无数据，跳过")
            continue

        csv_path = DATA / f"标签结果_{prefix}.csv"
        xlsx_path = DATA / f"标签结果_{prefix}.xlsx"

        # ── 写 CSV ──
        with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction='ignore')
            writer.writeheader()
            writer.writerows(rows)
        print(f"  📂 CSV: {csv_path.name} ({len(rows)} 条)")

        # ── 写 Excel（两个 sheet）──
        from openpyxl import Workbook
        wb = Workbook()

        # Sheet 1: 汇总
        ws_summary = wb.active
        ws_summary.title = "汇总"
        write_summary_sheet(ws_summary, rows, label)

        # Sheet 2: 数据明细
        ws_data = wb.create_sheet("数据明细")
        # 写表头
        for ci, fn in enumerate(fieldnames):
            cell = ws_data.cell(row=1, column=ci+1, value=fn)
            cell.font = header_font; cell.fill = header_fill
            cell.alignment = center; cell.border = thin_border
        # 写数据
        for ri, r in enumerate(rows):
            for ci, fn in enumerate(fieldnames):
                cell = ws_data.cell(row=ri+2, column=ci+1, value=r.get(fn, ""))
                cell.font = body_font; cell.border = thin_border
                cell.alignment = left_align if ci > 3 else center
        ws_data.freeze_panes = "A2"
        # 列宽
        for ci in range(len(fieldnames)):
            ws_data.column_dimensions[get_column_letter(ci+1)].width = 12

        wb.save(xlsx_path)
        print(f"  📂 Excel: {xlsx_path.name} (汇总 + 数据明细)")

    print(f"\n✅ 完成")


if __name__ == "__main__":
    main()
