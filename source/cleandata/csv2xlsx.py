"""CSV → Excel 转换工具"""
import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys

csv_path = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("res/data/标签结果.csv")
excel_path = csv_path.with_suffix(".xlsx")

with open(csv_path, "r", encoding="utf-8-sig") as f:
    reader = csv.DictReader(f)
    fieldnames = reader.fieldnames
    rows = list(reader)

print(f"读取 {len(rows)} 行，{len(fieldnames)} 列: {fieldnames}")

wb = Workbook()
ws = wb.active
ws.title = "标签结果"

# 样式
header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
data_font = Font(name="微软雅黑", size=10)
data_align = Alignment(vertical="top", wrap_text=True)
link_font = Font(name="微软雅黑", size=10, color="0563C1", underline="single")
number_align = Alignment(horizontal="center", vertical="top")

# 写表头
for col_idx, name in enumerate(fieldnames, 1):
    cell = ws.cell(row=1, column=col_idx, value=name)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

# 写数据
for row_idx, row_data in enumerate(rows, 2):
    for col_idx, name in enumerate(fieldnames, 1):
        value = row_data.get(name, "")
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = data_font
        cell.border = thin_border
        if name == "链接":
            cell.font = link_font
            if value.startswith("http"):
                cell.hyperlink = value
            cell.alignment = data_align
        elif name in ("点赞数", "评论数"):
            try:
                cell.value = int(value)
            except ValueError:
                pass
            cell.alignment = number_align
        else:
            cell.alignment = data_align

# 列宽
col_widths = {
    "链接": 35, "标签": 30, "标题": 45, "正文开头": 50,
    "点赞数": 10, "评论数": 10, "AI证据理由": 55, "AI原始JSON": 45,
}
for col_idx, name in enumerate(fieldnames, 1):
    ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(name, 15)

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(fieldnames))}{len(rows) + 1}"
wb.save(excel_path)
print(f"✅ Excel 已保存: {excel_path}")
