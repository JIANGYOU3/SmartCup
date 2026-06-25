"""清洗爬取结果 — 删除污染数据"""
import csv, re
from pathlib import Path

INPUT = "res/data/zhihu/output/爬取结果.csv"
OUTPUT = "res/data/zhihu/output/爬取结果_清洗后.csv"

NSFW = ["飞机杯", "成人用品", "TENGA", "自慰", "充气娃娃", "振动棒", "按摩棒"]

with open(INPUT, "r", encoding="utf-8-sig") as f:
    rows = list(csv.DictReader(f))

total = len(rows)
kept = []
removed_reasons = {}

for r in rows:
    content = r.get("回答内容", "").strip()
    title = r.get("问题标题", "").strip()
    post_type = r.get("帖子类型", "")
    full_text = f"{title} {content}"

    # 1. 视频 — 删除
    if post_type == "视频":
        removed_reasons["视频无内容"] = removed_reasons.get("视频无内容", 0) + 1
        continue

    # 2. 空内容 — 删除
    if not content and not title:
        removed_reasons["空内容"] = removed_reasons.get("空内容", 0) + 1
        continue

    # 3. NSFW — 删除
    hit = None
    for kw in NSFW:
        if kw in full_text:
            hit = kw; break
    if hit:
        removed_reasons[f"NSFW-{hit}"] = removed_reasons.get(f"NSFW-{hit}", 0) + 1
        continue

    # 4. 太短（<20字）且无赞无评 — 删除
    if len(content) < 20 and int(r.get("赞同数", 0) or 0) == 0 and int(r.get("评论数", 0) or 0) == 0:
        removed_reasons["短内容无互动"] = removed_reasons.get("短内容无互动", 0) + 1
        continue

    kept.append(r)

# 写入
with open(OUTPUT, "w", encoding="utf-8-sig", newline="") as f:
    w = csv.DictWriter(f, fieldnames=rows[0].keys())
    w.writeheader()
    w.writerows(kept)

print(f"原始: {total} 条")
print(f"保留: {len(kept)} 条")
print(f"删除: {total - len(kept)} 条")
for reason, cnt in sorted(removed_reasons.items(), key=lambda x: -x[1]):
    print(f"  ❌ {reason}: {cnt}")
print(f"\n✅ 清洗后: {OUTPUT}")

# 转 Excel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

excel_path = Path(OUTPUT).with_suffix(".xlsx")
wb = Workbook()
ws = wb.active
ws.title = "清洗后数据"

hf = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
b = Border(left=Side(style="thin", color="D9D9D9"), right=Side(style="thin", color="D9D9D9"),
           top=Side(style="thin", color="D9D9D9"), bottom=Side(style="thin", color="D9D9D9"))
df = Font(name="微软雅黑", size=10)
da = Alignment(vertical="top", wrap_text=True)
lf = Font(name="微软雅黑", size=10, color="0563C1", underline="single")
na = Alignment(horizontal="center", vertical="top")

fieldnames = list(rows[0].keys())
for ci, name in enumerate(fieldnames, 1):
    c = ws.cell(row=1, column=ci, value=name)
    c.font = hf; c.fill = hfill; c.alignment = ha; c.border = b

for ri, row in enumerate(kept, 2):
    for ci, name in enumerate(fieldnames, 1):
        v = row.get(name, "")
        c = ws.cell(row=ri, column=ci, value=v)
        c.font = df; c.border = b
        if name == "问答链接":
            c.font = lf
            if str(v).startswith("http"): c.hyperlink = str(v)
            c.alignment = da
        elif name in ("赞同数","评论数","问题被浏览次数","问题回答个数","问题评论个数"):
            try: c.value = int(v)
            except: pass
            c.alignment = na
        else:
            c.alignment = da

widths = {"关键词":25, "帖子类型":10, "问题被浏览次数":14, "问题回答个数":14,
          "问题评论个数":14, "问题标题":40, "问题内容":50, "答主昵称":14,
          "回答时间":16, "赞同数":10, "评论数":10, "回答内容":55, "问答链接":35}
for ci, name in enumerate(fieldnames, 1):
    ws.column_dimensions[get_column_letter(ci)].width = widths.get(name, 15)

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(fieldnames))}{len(kept)+1}"
wb.save(excel_path)
print(f"✅ Excel: {excel_path}")
