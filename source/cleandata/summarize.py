"""
数据汇总 — 生成 Excel 汇总表（写在 标签结果_最终.xlsx 的"汇总"sheet）

用法：
  conda run -n SmartCup python source/cleandata/summarize.py
"""

import csv
import re
import json
import sys
from pathlib import Path
from collections import Counter, defaultdict
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))

DATA = Path(__file__).resolve().parent.parent.parent / "res/data/douyin/output"
INPUT = DATA / "标签结果_最终.csv"
EXCEL = DATA / "标签结果_最终.xlsx"


def safe_int(v):
    try: return int(v or 0)
    except: return 0


def main():
    # ── 读取数据 ──
    rows = []
    with open(INPUT, encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            rows.append(r)

    N = len(rows)
    total_likes = sum(safe_int(r.get("点赞数", 0)) for r in rows)
    total_comments = sum(safe_int(r.get("评论数", 0)) for r in rows)
    total_collects = sum(safe_int(r.get("收藏数", 0)) for r in rows)
    total_shares = sum(safe_int(r.get("分享数", 0)) for r in rows)
    total_followers = sum(safe_int(r.get("作者粉丝量", 0)) for r in rows)

    # ── 统计各维度 ──
    label_count = Counter()
    emotion_count = Counter()
    topic_count = Counter()
    relevance_count = Counter()
    value_count = Counter()
    pain_points = Counter()
    all_keywords = Counter()

    for r in rows:
        for label in r.get("内容标签", "").replace("，", ",").split(","):
            label = label.strip().strip("[]'\" ")
            if label:
                label_count[label] += 1
        emo = r.get("用户情绪", "").strip()
        if emo: emotion_count[emo] += 1
        topic = r.get("话题类别", "").strip()
        if topic: topic_count[topic] += 1
        rel = r.get("相关性等级", "").strip()
        if rel: relevance_count[rel] += 1
        val = r.get("数据价值等级", "").strip()
        if val: value_count[val] += 1
        pains = r.get("需求痛点", "").strip()
        if pains and pains != "无":
            for p in pains.replace("；", ";").replace("，", ",").split(";"):
                p = p.strip()
                if p and len(p) > 2:
                    pain_points[p[:40]] += 1
        kws = r.get("关键词", "").strip()
        if kws:
            for kw in kws.replace("；", ";").replace("，", ",").split(","):
                kw = kw.strip().strip("[]'\" ")
                if kw:
                    all_keywords[kw] += 1

    # ── 写入 Excel 汇总 sheet ──
    wb = load_workbook(EXCEL)
    if "汇总" in wb.sheetnames:
        del wb["汇总"]
    ws = wb.create_sheet("汇总", 0)

    # 样式
    title_font = Font(name="微软雅黑", size=14, bold=True, color="1F4E79")
    h2_font = Font(name="微软雅黑", size=12, bold=True, color="2E75B6")
    header_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2E75B6")
    body_font = Font(name="微软雅黑", size=10)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    center = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    COL_W = {"A": 6, "B": 32, "C": 16, "D": 16, "E": 16}
    for col, w in COL_W.items():
        ws.column_dimensions[col].width = w

    row = 1

    def write_title(text, r):
        nonlocal row
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        cell = ws.cell(row=r, column=1, value=text)
        cell.font = title_font
        return r + 2

    def write_h2(text, r):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        cell = ws.cell(row=r, column=1, value=text)
        cell.font = h2_font
        return r + 1

    def write_table(r, headers, data, col_widths=None):
        """写表格，headers=[(col, width, label), ...], data=[[col1, col2, ...], ...]"""
        for ci, (col, width, label) in enumerate(headers):
            cell = ws.cell(row=r, column=ci+1, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin_border
        r += 1
        for dr in data:
            for ci, val in enumerate(dr):
                cell = ws.cell(row=r, column=ci+1, value=val)
                cell.font = body_font
                cell.alignment = left_align if ci == 1 else center
                cell.border = thin_border
            r += 1
        return r + 1

    # ═══════════════════ 总览 ═══════════════════
    row = write_title("📊 抖音智能水杯数据汇总", row)
    row = write_table(row,
        [("A", 6, "#"), ("B", 32, "指标"), ("C", 16, "数值"), ("D", 16, "占比"), ("E", 16, "备注")],
        [
            [1, "视频总数", f"{N:,}", "100%", f"文件: {EXCEL.name}"],
            [2, "总点赞数", f"{total_likes:,}", "-", f"均值 {total_likes//N:,}/条"],
            [3, "总评论数", f"{total_comments:,}", "-", f"均值 {total_comments//N:,}/条"],
            [4, "总收藏数", f"{total_collects:,}", "-", f"均值 {total_collects//N:,}/条"],
            [5, "总分享数", f"{total_shares:,}", "-", f"均值 {total_shares//N:,}/条"],
            [6, "作者粉丝总量", f"{total_followers:,}", "-", "-"],
        ]
    )

    # ═══════════════════ 内容标签 ═══════════════════
    row = write_h2("🏷️ 内容标签分布", row)
    top_labels = label_count.most_common(20)
    row = write_table(row,
        [("A", 6, "#"), ("B", 32, "内容标签"), ("C", 16, "数量"), ("D", 16, "占比"), ("E", 16, "")],
        [[i+1, label, cnt, f"{cnt/N*100:.1f}%", ""] for i, (label, cnt) in enumerate(top_labels)]
    )

    # ═══════════════════ 话题+情绪+相关性+价值 ═══════════════════
    row = write_h2("📂 话题类别", row)
    data = [[i+1, t, c, f"{c/N*100:.1f}%", ""] for i, (t, c) in enumerate(topic_count.most_common(15))]
    row = write_table(row, [("A", 6, "#"), ("B", 32, "话题"), ("C", 16, "数量"), ("D", 16, "占比"), ("E", 16, "")], data)

    row = write_h2("😊 用户情绪", row)
    data = [[i+1, e, c, f"{c/N*100:.1f}%", ""] for i, (e, c) in enumerate(emotion_count.most_common())]
    row = write_table(row, [("A", 6, "#"), ("B", 32, "情绪"), ("C", 16, "数量"), ("D", 16, "占比"), ("E", 16, "")], data)

    row = write_h2("🎯 相关性", row)
    data = [[i+1, rk, c, f"{c/N*100:.1f}%", ""] for i, (rk, c) in enumerate(relevance_count.most_common())]
    row = write_table(row, [("A", 6, "#"), ("B", 32, "等级"), ("C", 16, "数量"), ("D", 16, "占比"), ("E", 16, "")], data)

    row = write_h2("💎 数据价值", row)
    data = [[i+1, v, c, f"{c/N*100:.1f}%", ""] for i, (v, c) in enumerate(value_count.most_common())]
    row = write_table(row, [("A", 6, "#"), ("B", 32, "等级"), ("C", 16, "数量"), ("D", 16, "占比"), ("E", 16, "")], data)

    # ═══════════════════ 痛点 ═══════════════════
    row = write_h2("🔍 用户需求痛点 Top 20", row)
    top_pain = pain_points.most_common(20)
    row = write_table(row,
        [("A", 6, "#"), ("B", 32, "痛点"), ("C", 16, "提及次数"), ("D", 16, ""), ("E", 16, "")],
        [[i+1, p[:50], c, "", ""] for i, (p, c) in enumerate(top_pain)]
    )

    # ═══════════════════ 高频关键词 ═══════════════════
    row = write_h2("🔑 AI 提取关键词 Top 25", row)
    top_kw = all_keywords.most_common(25)
    # 每行5个
    data = []
    for i in range(0, len(top_kw), 5):
        chunk = top_kw[i:i+5]
        data.append([f"{kw}({c})" for kw, c in chunk])
    headers = [("A", 10, "关键词"), ("B", 10, ""), ("C", 10, ""), ("D", 10, ""), ("E", 10, "")]
    for ci, (col, w, label) in enumerate(headers):
        ws.column_dimensions[col].width = max(w, 22)
    for ci in range(5):
        cell = ws.cell(row=row, column=ci+1, value=f"#{ci+1}")
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border
    row += 1
    for dr in data:
        for ci, val in enumerate(dr):
            cell = ws.cell(row=row, column=ci+1, value=val)
            cell.font = body_font
            cell.alignment = center
            cell.border = thin_border
        row += 1

    # 冻结首行
    ws.freeze_panes = "A2"
    wb.save(EXCEL)
    print(f"✅ 汇总已写入: {EXCEL} → 汇总 sheet")
    print(f"   视频: {N:,} 条 | 👍{total_likes:,} | 💬{total_comments:,} | ⭐{total_collects:,}")


if __name__ == "__main__":
    main()
