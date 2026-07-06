"""
爬取高评论帖子的评论区内容 + 按评论数重排数据

流程：
  1. 读取清洗后的爬取结果
  2. 筛出评论数 > 50 的帖子
  3. 调用知乎 API 爬取每条评论（含回复）
  4. 全部数据按评论数降序排列
  5. 评论区一条评论占一格，追加到对应行后面
  6. 输出 CSV + Excel
"""

import csv, re, time, random, json, os
from pathlib import Path
from collections import deque

import requests
from tqdm import tqdm

from source.common.paths import load_env, get_project_root

load_env()
COOKIE = os.getenv("ZHIHU_COOKIE", "")

INPUT = "res/data/zhihu/output/爬取结果_清洗后.csv"
OUTPUT = "res/data/zhihu/output/爬取结果_含评论.csv"

LIMIT_PER_PAGE = 20  # API 每页条数

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Cookie": COOKIE,
    "x-requested-with": "fetch",
    "x-api-version": "3.0.40",
}


def fetch_comments_answer(aid: str, max_count: int) -> list[str]:
    """爬取回答的评论（含回复）"""
    all_comments = []
    offset = 0
    session = requests.Session()
    session.headers.update(HEADERS)

    while len(all_comments) < max_count:
        url = f"https://www.zhihu.com/api/v4/answers/{aid}/comments?order=normal&limit={LIMIT_PER_PAGE}&offset={offset}"
        time.sleep(1.5 + random.uniform(0, 1))
        resp = session.get(url, timeout=30)

        if resp.status_code != 200:
            print(f"    ⚠️ HTTP {resp.status_code}")
            break

        data = resp.json()
        items = data.get("data", [])
        if not items:
            break

        for item in items:
            content = re.sub(r'<[^>]+>', '', item.get("content", "")).strip()
            author = item.get("author", {}).get("member", {}).get("name", "")
            if not author:
                author = item.get("author", {}).get("name", "")

            # 回复
            replies = item.get("child_comments", [])
            reply_texts = []
            for rep in replies[:20]:  # 只取前20条回复
                r_content = re.sub(r'<[^>]+>', '', rep.get("content", "")).strip()
                r_author = rep.get("author", {}).get("member", {}).get("name", "")
                if not r_author:
                    r_author = rep.get("author", {}).get("name", "")
                reply_texts.append(f"  ↳ [{r_author}] {r_content}")

            line = f"[{author}] {content}"
            if reply_texts:
                line += "\n" + "\n".join(reply_texts)
            all_comments.append(line)

        offset += LIMIT_PER_PAGE
        if data.get("paging", {}).get("is_end"):
            break

    return all_comments


def fetch_comments_article(pid: str, max_count: int) -> list[str]:
    """专栏文章评论 — 知乎已封禁API，暂不支持"""
    print(f"    ⚠️ 专栏文章评论API已封禁，跳过")
    return []


def extract_id(url: str) -> tuple:
    """提取 question_id, answer_id, 或 post_id"""
    m = re.search(r'/question/(\d+)/answer/(\d+)', url)
    if m: return ('answer', m.group(2), m.group(1))
    m = re.search(r'/question/(\d+)', url)
    if m: return ('question', m.group(1), None)
    m = re.search(r'/p/(\d+)', url)
    if m: return ('article', m.group(1), None)
    return (None, None, None)


print("=" * 60)
print("评论区爬取 + 数据重排")
print("=" * 60)

# ── 1. 读取数据 ──
with open(INPUT, "r", encoding="utf-8-sig") as f:
    rows = list(csv.DictReader(f))
print(f"\n📂 读取: {len(rows)} 条")

# ── 2. 筛出需爬评论的帖子（问答帖 + 评论>0）──
to_crawl = []
for r in rows:
    c = int(r.get("评论数", 0) or 0)
    ptype = r.get("帖子类型", "")
    url = r.get("问答链接", "")
    post_type, target_id, _ = extract_id(url)
    if post_type == "answer" and c > 0:
        to_crawl.append(r)

# 统计
article_skip = sum(1 for r in rows if r.get("帖子类型") == "专栏文章" and int(r.get("评论数",0) or 0) > 0)
zero_cmt = sum(1 for r in rows if int(r.get("评论数",0) or 0) == 0)
print(f"🔍 需爬评论: {len(to_crawl)} 条问答帖（评论>0）")
print(f"⏭️ 跳过专栏文章: {article_skip} 条（API封禁）")
print(f"⏭️ 零评论帖子: {zero_cmt} 条")

if not to_crawl:
    print("⚠️ 无需爬取，仅做排序")
else:
    print(f"\n🐌 开始爬取评论（约 {len(to_crawl)} 条帖子，每条 2-3s）...")
    print()
    comment_data = {}  # link -> [comment1_str, comment2_str, ...]

    pbar = tqdm(total=len(to_crawl), desc="爬评论", unit="帖")
    for r in to_crawl:
        url = r["问答链接"]
        post_type, target_id, _ = extract_id(url)
        max_c = int(r.get("评论数", 0) or 0)

        comments = fetch_comments_answer(target_id, max_c)
        comment_data[url] = comments
        pbar.set_postfix_str(f"评{max_c}→得{len(comments)}")
        pbar.update(1)
    pbar.close()

# ── 4. 按评论数降序排列 ──
rows.sort(key=lambda r: -int(r.get("评论数", 0) or 0))
print(f"\n📊 已按评论数降序排列")

# ── 5. 确定最多评论列数 ──
max_cols = 0
for comments in comment_data.values():
    max_cols = max(max_cols, len(comments))
print(f"   最多评论列数: {max_cols}")

# ── 6. 输出 ──
base_fields = list(rows[0].keys())
comment_fields = [f"评论{i+1}" for i in range(max_cols)] if max_cols > 0 else []
all_fields = base_fields + comment_fields

with open(OUTPUT, "w", encoding="utf-8-sig", newline="") as f:
    writer = csv.DictWriter(f, fieldnames=all_fields, extrasaction="ignore")
    writer.writeheader()

    for r in rows:
        row = dict(r)
        url = r["问答链接"]
        comments = comment_data.get(url, [])
        for i in range(max_cols):
            row[f"评论{i+1}"] = comments[i] if i < len(comments) else ""
        writer.writerow(row)

print(f"\n✅ CSV: {OUTPUT}")

# ── 7. 转 Excel ──
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

excel_path = Path(OUTPUT).with_suffix(".xlsx")
wb = Workbook()
ws = wb.active
ws.title = "含评论数据"

hf = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
b = Border(left=Side(style="thin", color="D9D9D9"), right=Side(style="thin", color="D9D9D9"),
           top=Side(style="thin", color="D9D9D9"), bottom=Side(style="thin", color="D9D9D9"))
df = Font(name="微软雅黑", size=10)
da = Alignment(vertical="top", wrap_text=True)
lf = Font(name="微软雅黑", size=10, color="0563C1", underline="single")
na = Alignment(horizontal="center", vertical="top")

for ci, name in enumerate(all_fields, 1):
    c = ws.cell(row=1, column=ci, value=name)
    c.font = hf; c.fill = hfill; c.alignment = ha; c.border = b

for ri, row in enumerate(rows, 2):
    url = row.get("问答链接", "")
    comments = comment_data.get(url, [])
    for ci, name in enumerate(all_fields, 1):
        if name.startswith("评论") and name[2:].isdigit():
            # 从 comment_data 取，不依赖 row
            idx = int(name.replace("评论", "")) - 1
            v = comments[idx] if idx < len(comments) else ""
        else:
            v = row.get(name, "")
        c = ws.cell(row=ri, column=ci, value=v)
        c.font = df; c.border = b
        if name == "问答链接":
            c.font = lf
            if str(v).startswith("http"): c.hyperlink = str(v)
            c.alignment = da
        elif name in ("赞同数", "评论数", "问题被浏览次数", "问题回答个数", "问题评论个数"):
            try: c.value = int(v)
            except: pass
            c.alignment = na
        elif name.startswith("评论"):
            c.alignment = Alignment(vertical="top", wrap_text=True, horizontal="left")
        else:
            c.alignment = da

# 列宽
for ci, name in enumerate(all_fields, 1):
    if name.startswith("评论"):
        ws.column_dimensions[get_column_letter(ci)].width = 45
    elif name in ("回答内容", "问题内容"):
        ws.column_dimensions[get_column_letter(ci)].width = 55
    elif name == "问答链接":
        ws.column_dimensions[get_column_letter(ci)].width = 35
    elif name in ("赞同数", "评论数"):
        ws.column_dimensions[get_column_letter(ci)].width = 10
    else:
        ws.column_dimensions[get_column_letter(ci)].width = 20

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(all_fields))}{len(rows) + 1}"
wb.save(excel_path)
print(f"✅ Excel: {excel_path}")
print(f"\n{'='*60}")
print("完成！")
print(f"{'='*60}")
