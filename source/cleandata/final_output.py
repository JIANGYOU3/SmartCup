"""
最终输出 — 规范命名 + 文件夹归档 + 按时间分割 + 额外汇总（近三年/中高相关性/负面情绪）

用法：
  conda run -n SmartCup python source/cleandata/final_output.py
"""

import csv
import re
import sys
import json
from pathlib import Path
from datetime import datetime
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))

PROJECT = Path(__file__).resolve().parent.parent.parent
DATA = PROJECT / "res/data/douyin/output"
INPUT = DATA / "标签结果_最终.csv"
VERSION = "1.0"
TODAY = datetime.now().strftime("%Y%m%d")
OUT_DIR = DATA / f"数据清洗结果_{TODAY}"

CUTOFF = datetime(2023, 7, 1)

# ── 样式 ──
TITLE_FONT = Font(name="微软雅黑", size=14, bold=True, color="1F4E79")
H2_FONT = Font(name="微软雅黑", size=12, bold=True, color="2E75B6")
HEADER_FONT = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="2E75B6")
BODY_FONT = Font(name="微软雅黑", size=10)
NEG_FILL = PatternFill("solid", fgColor="FFF2CC")  # 负面高亮
THIN_BORDER = Border(left=Side(style="thin"), right=Side(style="thin"),
                      top=Side(style="thin"), bottom=Side(style="thin"))
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def safe_int(v):
    try: return int(v or 0)
    except: return 0


def parse_date(r):
    t = r.get("发布时间", "").strip()
    try: return datetime.strptime(t[:10], '%Y-%m-%d')
    except: return None


def write_summary_sheet(ws, rows, label):
    """写汇总 sheet"""
    N = len(rows)
    total_likes = sum(safe_int(r.get("点赞数", 0)) for r in rows)
    total_comments = sum(safe_int(r.get("评论数", 0)) for r in rows)
    total_collects = sum(safe_int(r.get("收藏数", 0)) for r in rows)
    total_shares = sum(safe_int(r.get("分享数", 0)) for r in rows)

    label_cnt = Counter()
    emotion_cnt = Counter()
    topic_cnt = Counter()
    value_cnt = Counter()
    relevance_cnt = Counter()
    pain_cnt = Counter()
    kw_cnt = Counter()

    for r in rows:
        for lb in r.get("内容标签", "").replace("，", ",").split(","):
            lb = lb.strip().strip("[]'\" ")
            if lb: label_cnt[lb] += 1
        emo = r.get("用户情绪", "").strip()
        if emo: emotion_cnt[emo] += 1
        tp = r.get("话题类别", "").strip()
        if tp: topic_cnt[tp] += 1
        vl = r.get("数据价值等级", "").strip()
        if vl: value_cnt[vl] += 1
        rl = r.get("相关性等级", "").strip()
        if rl: relevance_cnt[rl] += 1
        ps = r.get("需求痛点", "").strip()
        if ps and ps != "无":
            for p in ps.replace("；", ";").replace("，", ",").split(";"):
                p = p.strip()
                if p and len(p) > 2: pain_cnt[p[:50]] += 1
        ks = r.get("关键词", "").strip()
        if ks:
            for k in ks.replace("；", ";").replace("，", ",").split(","):
                k = k.strip().strip("[]'\" ")
                if k: kw_cnt[k] += 1

    for c, w in [("A", 6), ("B", 32), ("C", 16), ("D", 16), ("E", 16)]:
        ws.column_dimensions[c].width = w

    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    ws.cell(row=r, column=1, value=f"📊 抖音智能水杯数据汇总 — {label}").font = TITLE_FONT
    r += 2

    def section(title, items, top_n=20):
        nonlocal r
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        ws.cell(row=r, column=1, value=title).font = H2_FONT
        r += 1
        for ci, v in enumerate(["#", "名称", "数量", "占比", ""]):
            c = ws.cell(row=r, column=ci+1, value=v)
            c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = CENTER; c.border = THIN_BORDER
        r += 1
        for i, (name, cnt) in enumerate(items[:top_n]):
            for ci, v in enumerate([i+1, name, cnt, f"{cnt/max(N,1)*100:.1f}%", ""]):
                c = ws.cell(row=r, column=ci+1, value=v)
                c.font = BODY_FONT; c.alignment = LEFT if ci == 1 else CENTER; c.border = THIN_BORDER
            r += 1
        r += 1

    # 总览
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    ws.cell(row=r, column=1, value="📋 数据总览").font = H2_FONT
    r += 1
    overview = [
        ["#", "指标", "数值", "", ""],
        [1, "视频总数", f"{N:,}", "", ""],
        [2, "总点赞数", f"{total_likes:,}", f"均值 {total_likes//N:,}/条", ""],
        [3, "总评论数", f"{total_comments:,}", f"均值 {total_comments//N:,}/条", ""],
        [4, "总收藏数", f"{total_collects:,}", f"均值 {total_collects//N:,}/条", ""],
        [5, "总分享数", f"{total_shares:,}", f"均值 {total_shares//N:,}/条", ""],
    ]
    for dr in overview:
        for ci, v in enumerate(dr):
            c = ws.cell(row=r, column=ci+1, value=v)
            c.font = HEADER_FONT if dr == overview[0] else BODY_FONT
            c.fill = HEADER_FILL if dr == overview[0] else PatternFill()
            c.alignment = CENTER; c.border = THIN_BORDER
        r += 1
    r += 1

    section("🏷️ 内容标签分布", label_cnt.most_common(20))
    section("📂 话题类别", topic_cnt.most_common(15))
    section("😊 用户情绪", emotion_cnt.most_common())
    section("🎯 相关性等级", relevance_cnt.most_common())
    section("💎 数据价值等级", value_cnt.most_common())
    section("🔍 用户需求痛点 Top 20", pain_cnt.most_common(20))

    # 关键词
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    ws.cell(row=r, column=1, value="🔑 AI提取关键词 Top 25").font = H2_FONT
    r += 1
    for i in range(0, 25, 5):
        chunk = list(kw_cnt.most_common(25))[i:i+5]
        for ci, (kw, cnt) in enumerate(chunk):
            c = ws.cell(row=r, column=ci+1, value=f"{kw}({cnt})")
            c.font = BODY_FONT; c.alignment = CENTER; c.border = THIN_BORDER
        r += 1

    ws.freeze_panes = "A2"


def write_data_sheet(ws, rows, fieldnames):
    """写数据明细 sheet"""
    for ci, fn in enumerate(fieldnames):
        c = ws.cell(row=1, column=ci+1, value=fn)
        c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = CENTER; c.border = THIN_BORDER
    for ri, row in enumerate(rows):
        for ci, fn in enumerate(fieldnames):
            c = ws.cell(row=ri+2, column=ci+1, value=row.get(fn, ""))
            c.font = BODY_FONT; c.border = THIN_BORDER
    ws.freeze_panes = "A2"


def write_negative_sheet(ws, rows):
    """负面情绪专题 sheet"""
    neg = [r for r in rows if "负面" in r.get("用户情绪", "")]
    ws.title = "负面情绪"

    cols = ["视频标题", "视频文案", "用户情绪", "需求痛点", "一句话总结", "点赞数", "评论数", "视频链接"]
    for ci, fn in enumerate(cols):
        c = ws.cell(row=1, column=ci+1, value=fn)
        c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = CENTER; c.border = THIN_BORDER

    for ri, row in enumerate(neg):
        for ci, fn in enumerate(cols):
            c = ws.cell(row=ri+2, column=ci+1, value=row.get(fn, ""))
            c.font = BODY_FONT
            c.fill = NEG_FILL
            c.border = THIN_BORDER
            c.alignment = LEFT

    # 负面评论单独列出
    neg_comments = []
    for row in neg:
        for i in range(1, 21):
            text = row.get(f"评论{i}", "").strip()
            if text:
                neg_comments.append({
                    "视频标题": row.get("视频标题", "")[:60],
                    "评论内容": text,
                    "评论点赞": row.get(f"评论{i}点赞", "0"),
                    "评论用户": row.get(f"评论{i}用户", ""),
                    "视频链接": row.get("视频链接", ""),
                })

    if neg_comments:
        # 评论起始列
        start_col = len(cols) + 2
        comment_cols = ["视频标题", "评论内容", "评论点赞", "评论用户", "视频链接"]
        for ci, fn in enumerate(comment_cols):
            c = ws.cell(row=1, column=start_col+ci, value=f"💬{fn}")
            c.font = HEADER_FONT; c.fill = PatternFill("solid", fgColor="E2EFDA")
            c.alignment = CENTER; c.border = THIN_BORDER
        for ri, nc in enumerate(neg_comments):
            for ci, fn in enumerate(comment_cols):
                c = ws.cell(row=ri+2, column=start_col+ci, value=nc[fn])
                c.font = BODY_FONT; c.border = THIN_BORDER; c.alignment = LEFT

    ws.freeze_panes = "A2"

    # 列宽
    for c, w in [("A", 40), ("B", 40), ("C", 12), ("D", 30), ("E", 50), ("F", 10), ("G", 10), ("H", 35)]:
        ws.column_dimensions[c].width = w

    return len(neg), len(neg_comments)


def main():
    # ── 读取 ──
    with open(INPUT, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        all_rows = list(reader)
        fieldnames = reader.fieldnames

    print(f"📂 读取: {len(all_rows)} 条, {len(fieldnames)} 列")

    # ── 创建输出目录 ──
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    print(f"📁 输出目录: {OUT_DIR}")

    # ── 按时间拆分 ──
    recent, older = [], []
    for r in all_rows:
        dt = parse_date(r)
        if dt and dt >= CUTOFF:
            recent.append(r)
        else:
            older.append(r)

    # ── 近三年 + 中高相关性（用于额外汇总）──
    recent_high_rel = [r for r in recent if r.get("相关性等级", "") in ("高", "中")]

    print(f"  近三年: {len(recent)} 条")
    print(f"  三年前: {len(older)} 条")
    print(f"  近三年+中高相关性: {len(recent_high_rel)} 条")

    # ═══════════════════════════════════════════
    #  生成文件
    # ═══════════════════════════════════════════

    for label, rows, suffix in [
        ("近三年", recent, "近三年"),
        ("三年前", older, "三年前"),
    ]:
        if not rows:
            continue

        v_count = len(rows)
        c_count = sum(1 for r in rows for i in range(1, 21) if r.get(f"评论{i}", "").strip())
        csv_name = f"数据清洗结果_v{VERSION}_{suffix}_{v_count}条_{c_count}评_{TODAY}.csv"
        xlsx_name = f"数据清洗结果_v{VERSION}_{suffix}_{v_count}条_{c_count}评_{TODAY}.xlsx"

        # CSV
        csv_path = OUT_DIR / csv_name
        with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction='ignore')
            writer.writeheader()
            writer.writerows(rows)

        # Excel
        xlsx_path = OUT_DIR / xlsx_name
        wb = Workbook()

        ws1 = wb.active
        ws1.title = "汇总"
        write_summary_sheet(ws1, rows, f"{label} (v{VERSION})")

        ws2 = wb.create_sheet("数据明细")
        write_data_sheet(ws2, rows, fieldnames)

        wb.save(xlsx_path)
        print(f"  ✅ {xlsx_name} ({len(rows)} 条)")

    # ═══════════════════════════════════════════
    #  额外汇总：近三年 + 中高相关性 + 负面情绪
    # ═══════════════════════════════════════════
    special_name = f"数据清洗结果_v{VERSION}_近三年_中高相关性_含负面_{TODAY}.xlsx"
    special_path = OUT_DIR / special_name
    wb = Workbook()

    # Sheet 1: 汇总
    ws1 = wb.active
    ws1.title = "汇总"
    write_summary_sheet(ws1, recent_high_rel, f"近三年·中高相关性 (v{VERSION})")

    # Sheet 2: 数据明细
    ws2 = wb.create_sheet("数据明细")
    write_data_sheet(ws2, recent_high_rel, fieldnames)

    # Sheet 3: 负面情绪 + 负面评论
    ws3 = wb.create_sheet()
    neg_count, neg_comment_count = write_negative_sheet(ws3, recent_high_rel)

    wb.save(special_path)
    print(f"  ✅ {special_name} ({len(recent_high_rel)} 条, 负面 {neg_count} 条, 负面评论 {neg_comment_count} 条)")

    # ═══════════════════════════════════════════
    #  全量文件也放进去
    # ═══════════════════════════════════════════
    full_v = len(all_rows)
    full_c = sum(1 for r in all_rows for i in range(1, 21) if r.get(f"评论{i}", "").strip())
    full_name = f"数据清洗结果_v{VERSION}_全量_{full_v}条_{full_c}评_{TODAY}.xlsx"
    full_path = OUT_DIR / full_name
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "汇总"
    write_summary_sheet(ws1, all_rows, f"全量 (v{VERSION})")
    ws2 = wb.create_sheet("数据明细")
    write_data_sheet(ws2, all_rows, fieldnames)
    wb.save(full_path)

    full_csv = OUT_DIR / f"数据清洗结果_v{VERSION}_全量_{full_v}条_{full_c}评_{TODAY}.csv"
    with open(full_csv, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction='ignore')
        writer.writeheader()
        writer.writerows(all_rows)

    print(f"  ✅ 全量文件 ({len(all_rows)} 条)")

    # ── 最终统计 ──
    files = sorted(OUT_DIR.glob("*"))
    print(f"\n{'='*60}")
    print(f"  📁 {OUT_DIR.name}/")
    for f in files:
        size = f.stat().st_size
        size_str = f"{size/1024:.0f}KB" if size < 1024*1024 else f"{size/1024/1024:.1f}MB"
        print(f"    {f.name}  ({size_str})")
    print(f"{'='*60}")
    print(f"  ✅ 全部完成，共 {len(files)} 个文件")


if __name__ == "__main__":
    main()
