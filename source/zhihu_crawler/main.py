"""
知乎爬虫 — 主入口

用法：
  # 注意：-u 禁用输出缓冲，确保进度条正常显示
  conda run -n SmartCup python -u -m source.zhihu_crawler.main

  # 指定输入输出
  conda run -n SmartCup python -u -m source.zhihu_crawler.main \
    --input res/data/zhihu/raw/标题筛选版1.0.csv \
    --output res/data/zhihu/output/爬取结果.csv
"""

import sys
import argparse
from pathlib import Path

from .spiders.zhihu import run_batch_scrape


def csv_to_excel(csv_path: Path) -> Path:
    """将 CSV 转换为格式化的 Excel 文件"""
    import csv
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    excel_path = csv_path.with_suffix(".xlsx")

    with open(csv_path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames
        rows = list(reader)

    wb = Workbook()
    ws = wb.active
    ws.title = "爬取结果"

    header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    data_font = Font(name="微软雅黑", size=10)
    data_align = Alignment(vertical="top", wrap_text=True)
    link_font = Font(name="微软雅黑", size=10, color="0563C1", underline="single")
    number_align = Alignment(horizontal="center", vertical="top")

    for col_idx, name in enumerate(fieldnames, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, name in enumerate(fieldnames, 1):
            value = row_data.get(name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.border = thin_border
            if name == "问答链接":
                cell.font = link_font
                if str(value).startswith("http"):
                    cell.hyperlink = str(value)
                cell.alignment = data_align
            elif name in ("赞同数", "评论数", "问题被浏览次数", "问题回答个数", "问题评论个数"):
                try: cell.value = int(value)
                except ValueError: pass
                cell.alignment = number_align
            else:
                cell.alignment = data_align

    col_widths = {
        "关键词": 25, "帖子类型": 10, "问题被浏览次数": 14, "问题回答个数": 14,
        "问题评论个数": 14, "问题标题": 40, "问题内容": 50, "答主昵称": 14,
        "回答时间": 16, "赞同数": 10, "评论数": 10, "回答内容": 55, "问答链接": 35,
    }
    for col_idx, name in enumerate(fieldnames, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(name, 15)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(fieldnames))}{len(rows) + 1}"
    wb.save(excel_path)
    return excel_path


def main():
    parser = argparse.ArgumentParser(description="知乎内容批量爬虫")
    parser.add_argument("--input", "-i", default="res/data/zhihu/raw/标题筛选版1.0.csv")
    parser.add_argument("--output", "-o", default="res/data/zhihu/output/爬取结果.csv")
    parser.add_argument("--resume", action="store_true", default=True)
    parser.add_argument("--no-resume", action="store_true")
    args = parser.parse_args()

    project_root = Path(__file__).resolve().parent.parent.parent
    input_path = project_root / args.input
    output_path = project_root / args.output
    use_resume = not args.no_resume

    print("=" * 60, flush=True)
    print("知乎内容批量爬虫", flush=True)
    print("=" * 60, flush=True)
    print(f"  输入: {input_path}", flush=True)
    print(f"  输出: {output_path}", flush=True)
    print(f"  断点续爬: {'开启' if use_resume else '关闭'}", flush=True)
    print(flush=True)

    run_batch_scrape(
        input_csv=input_path,
        output_csv=output_path,
        resume=use_resume,
    )

    # ── 自动转 Excel ──
    if output_path.exists():
        print(f"\n📊 正在生成 Excel...", flush=True)
        try:
            excel_path = csv_to_excel(output_path)
            print(f"✅ Excel 已保存: {excel_path}", flush=True)
        except Exception as e:
            print(f"⚠️ Excel 生成失败: {e}", flush=True)


if __name__ == "__main__":
    main()
