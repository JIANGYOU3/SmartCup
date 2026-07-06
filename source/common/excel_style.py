"""Excel 样式 —— 统一的格式化输出"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 样式常量 ──────────────────────────────────────

HEADER_FONT = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

DATA_FONT = Font(name="微软雅黑", size=10)
DATA_ALIGN = Alignment(vertical="top", wrap_text=True)
LINK_FONT = Font(name="微软雅黑", size=10, color="0563C1", underline="single")
NUMBER_ALIGN = Alignment(horizontal="center", vertical="top")


def apply_excel_style(
    ws,
    rows: list[dict],
    fieldnames: list[str],
    *,
    sheet_title: str = "Sheet1",
    column_widths: dict[str, int] = None,
    link_columns: set[str] = None,
    number_columns: set[str] = None,
):
    """
    将数据写入 worksheet 并应用统一的蓝色表头样式。

    参数:
        ws: openpyxl Worksheet
        rows: 数据行列表
        fieldnames: 列名列表
        sheet_title: 工作表名称
        column_widths: {列名: 宽度} 映射
        link_columns: 需要超链接样式的列名集合
        number_columns: 需要数字居中对齐的列名集合
    """
    ws.title = sheet_title
    link_columns = link_columns or set()
    number_columns = number_columns or set()
    column_widths = column_widths or {}

    # ── 写表头 ──
    for col_idx, name in enumerate(fieldnames, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    # ── 写数据 ──
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, name in enumerate(fieldnames, 1):
            value = row_data.get(name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER

            if name in link_columns:
                cell.font = LINK_FONT
                if str(value).startswith("http"):
                    cell.hyperlink = str(value)
                cell.alignment = DATA_ALIGN
            elif name in number_columns:
                try:
                    cell.value = int(value)
                except (ValueError, TypeError):
                    pass
                cell.alignment = NUMBER_ALIGN
            else:
                cell.alignment = DATA_ALIGN

    # ── 列宽 ──
    for col_idx, name in enumerate(fieldnames, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = column_widths.get(name, 15)

    # ── 冻结表头 + 自动筛选 ──
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(fieldnames))}{len(rows) + 1}"


def csv_to_excel(
    csv_path: Path,
    excel_path: Path = None,
    *,
    sheet_title: str = "Sheet1",
    column_widths: dict[str, int] = None,
    link_columns: set[str] = None,
    number_columns: set[str] = None,
) -> Path:
    """
    读取 CSV 并输出格式化的 Excel 文件。

    返回生成的 .xlsx 文件路径。
    """
    import csv

    excel_path = excel_path or csv_path.with_suffix(".xlsx")

    with open(csv_path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames
        rows = list(reader)

    wb = Workbook()
    ws = wb.active
    apply_excel_style(
        ws, rows, fieldnames,
        sheet_title=sheet_title,
        column_widths=column_widths,
        link_columns=link_columns,
        number_columns=number_columns,
    )
    wb.save(excel_path)
    return excel_path
